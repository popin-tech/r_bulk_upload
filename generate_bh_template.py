from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font
import os

output_path = 'static/bh_import_template.xlsx'

# Create Workbook
wb = Workbook()
ws = wb.active
ws.title = "Import Template"

# Headers
headers = ['平台', 'AccID', 'R代理', '名稱', 'Budget', 'StartDate', 'EndDate', 'CPCGoal', 'CPAGoal', 'R的cv定義', 'Token']
ws.append(headers)

# Style Headers
for cell in ws[1]:
    cell.font = Font(bold=True)

# Sample Data
data = [
    # R Platform example
    ['R', 111222, '4A', 'Rixbee 範例帳戶', 50000, '2024-01-01', '2024-12-31', 15, 250, 'CompleteCheckout', ''],
    # D Platform example
    ['D', 333444, '', 'Discovery 範例帳戶', 30000, '2024-02-01', '2024-06-30', 10, 200, '', 'example_token_abcdef123']
]

for row in data:
    ws.append(row)

# --- Formatting ---
# Col B (AccID): Number format "0"
# Col F, G (Dates): Date format "yyyy-mm-dd" (Shifted by 1 due to inserted column)

# Set column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 10  # R代理
ws.column_dimensions['D'].width = 25  # Name
ws.column_dimensions['E'].width = 12  # Budget
ws.column_dimensions['F'].width = 12  # StartDate
ws.column_dimensions['G'].width = 12  # EndDate
ws.column_dimensions['H'].width = 10  # CPC
ws.column_dimensions['I'].width = 10  # CPA
ws.column_dimensions['J'].width = 25  # CV Def
ws.column_dimensions['K'].width = 30  # Token

# Apply formatting to all rows (1-1000)
for row in ws.iter_rows(min_row=2, max_row=1000):
    # AccID (Col 2)
    row[1].number_format = '0'
    # Dates (Col 6, 7) -> Index 5, 6
    row[5].number_format = 'yyyy-mm-dd'
    row[6].number_format = 'yyyy-mm-dd'

# --- Data Validation ---

# 1. Platform (Col A)
dv_platform = DataValidation(type="list", formula1='"R,D"', allow_blank=False)
dv_platform.error = '必須填寫 R 或 D'
dv_platform.errorTitle = '輸入錯誤'
ws.add_data_validation(dv_platform)
dv_platform.add('A2:A1000')

# 2. R Agent (Col C) - New
dv_agent = DataValidation(type="list", formula1='"4A,台客"', allow_blank=True)
dv_agent.error = '請選擇 4A 或 台客'
dv_agent.errorTitle = '輸入錯誤'
ws.add_data_validation(dv_agent)
dv_agent.add('C2:C1000')

# 3. CV Definition (Col J) -> Was I, now J
cv_options = [
    'CompleteCheckout', 
    'AddToCart', 
    'ViewContent', 
    'Checkout', 
    'Bookmark', 
    'Search', 
    'CompleteRegistration'
]
# Create formula string "Option1,Option2,..."
dv_formula = '"' + ','.join(cv_options) + '"'

dv_cv = DataValidation(type="list", formula1=dv_formula, allow_blank=True)
dv_cv.error = '請選擇清單中的項目'
dv_cv.errorTitle = '輸入無效'
dv_cv.prompt = '請從選單選擇轉換定義'
dv_cv.promptTitle = '選擇轉換定義'

ws.add_data_validation(dv_cv)
dv_cv.add('J2:J1000')

# Save
wb.save(output_path)
print(f"Generated {output_path} with data validation.")
