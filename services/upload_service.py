from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional
from datetime import datetime, timedelta


import unicodedata
import pandas as pd


# Localization Mappings
# Note: These are placeholder mappings based on standard practices.
DEVICE_MAP = {
    1: "Mobile",
    2: "PC",
    3: "TV",
    5: "Tablet",
    6: "loT",
    7: "TVBox"
}
TRAFFIC_MAP = {
    1: "App",
    2: "Web"
}
GENDER_MAP = {
    1: "男",
    2: "女",
    0: "未知"
}
AGE_MAP = {
    1: "18-24", 
    2: "25-34", 
    3: "35-44", 
    4: "45-54", 
    5: "55-64", 
    6: "65+"
}

PLATFORM_MAP = {
    1: "iOS",
    2: "Android", 3: "Others", 5: "MacOS", 6: "Windows", 101: "iPadOS"
}

BROWSER_MAP = {
    1: "Chrome", 2: "safari", 3: "Edge", 4: "FireFox", 5: "IE"
}

CATEGORY_MAP = {
    "IAB1": "藝術與娛樂活動", "IAB1-1": "閱讀書籍", "IAB1-2": "娛樂星聞", "IAB1-3": "藝術收藏",
    "IAB1-4": "喜劇", "IAB1-5": "影音串流", "IAB1-6": "音樂娛樂", "IAB1-7": "串流影集",
    "IAB2": "汽車與交通", "IAB2-1": "汽車配件", "IAB2-2": "汽車維修保養", "IAB2-3": "汽車買賣",
    "IAB2-4": "汽車愛好", "IAB2-5": "原廠認證中古車", "IAB2-6": "敞篷車", "IAB2-7": "轎跑車",
    "IAB2-8": "跨界休旅車", "IAB2-9": "柴油車", "IAB2-10": "電動車", "IAB2-11": "掀背車",
    "IAB2-12": "油電混合車", "IAB2-13": "豪華車", "IAB2-14": "小型廂型車", "IAB2-15": "機車",
    "IAB2-16": "越野車", "IAB2-17": "高性能跑車", "IAB2-18": "小型貨卡", "IAB2-19": "道路救援",
    "IAB2-20": "四門轎車", "IAB2-21": "商用車配件", "IAB2-22": "古董收藏車", "IAB2-23": "Wagon休旅車",
    "IAB3": "商業與產業", "IAB3-1": "廣告", "IAB3-2": "農業", "IAB3-3": "生物醫學",
    "IAB3-4": "商業軟體", "IAB3-5": "建築", "IAB3-6": "林業", "IAB3-7": "政府標案",
    "IAB3-8": "綠色能源", "IAB3-9": "人力資源", "IAB3-10": "物流", "IAB3-11": "行銷",
    "IAB3-12": "金屬產業", "IAB4": "求職職涯發展", "IAB4-1": "職涯規劃", "IAB4-2": "大學",
    "IAB4-3": "助學貸款", "IAB4-4": "招聘會", "IAB4-5": "求職", "IAB4-6": "履歷撰寫",
    "IAB4-7": "護理", "IAB4-8": "獎學金", "IAB4-9": "遠距工作", "IAB4-10": "軍事職涯",
    "IAB4-11": "職業諮詢", "IAB5": "教育學習", "IAB5-1": "國高中教育", "IAB5-2": "成人教育",
    "IAB5-3": "藝術史", "IAB5-4": "大學行政", "IAB5-5": "大學生活", "IAB5-6": "遠距學習",
    "IAB5-7": "線上英文學習", "IAB5-8": "語言學習", "IAB5-9": "研究所教育", "IAB5-10": "在家教育",
    "IAB5-11": "學習方法指導", "IAB5-12": "國小教師", "IAB5-13": "私立學校", "IAB5-14": "特殊教育",
    "IAB5-15": "學習障礙", "IAB6": "家庭與育兒", "IAB6-1": "收養", "IAB6-2": "嬰幼兒",
    "IAB6-3": "幼兒園", "IAB6-4": "親子教養", "IAB6-5": "學齡兒童教養", "IAB6-6": "青少年教養",
    "IAB6-7": "孕期照護", "IAB6-8": "特殊需求兒童", "IAB6-9": "銀髮照護", "IAB7": "健康與健身",
    "IAB7-1": "健身運動", "IAB7-2": "注意力缺失症", "IAB7-3": "愛滋病/HIV", "IAB7-4": "過敏疾病",
    "IAB7-5": "自然療法", "IAB7-6": "關節活動照護", "IAB7-7": "呼吸道健康", "IAB7-8": "特殊兒照護",
    "IAB7-9": "心理健康", "IAB7-10": "腦部照護", "IAB7-11": "腫瘤照護", "IAB7-12": "血脂管理",
    "IAB7-13": "慢性疲勞症候群", "IAB7-14": "長期疼痛照護", "IAB7-15": "呼吸道防護", "IAB7-16": "聽力障礙",
    "IAB7-17": "牙齒護理", "IAB7-18": "情緒健康", "IAB7-19": "皮膚照護", "IAB7-20": "血糖管理",
    "IAB7-21": "神經系統", "IAB7-22": "胃食道逆流症/胃酸逆流", "IAB7-23": "頭部舒緩", "IAB7-24": "心血管照護",
    "IAB7-25": "保健草藥", "IAB7-26": "身心靈療癒", "IAB7-27": "腸胃保健", "IAB7-28": "家庭暴力防治",
    "IAB7-29": "私密照護", "IAB7-30": "不孕症", "IAB7-31": "男性健康", "IAB7-32": "健康飲食管理",
    "IAB7-33": "骨骼健康", "IAB7-34": "情緒舒壓", "IAB7-35": "兒童醫學", "IAB7-36": "物理治療",
    "IAB7-37": "精神醫學", "IAB7-38": "銀髮健康", "IAB7-39": "親密關係", "IAB7-40": "睡眠障礙",
    "IAB7-41": "無菸生活", "IAB7-42": "物質濫用", "IAB7-43": "內分泌代謝", "IAB7-44": "體重管理",
    "IAB7-45": "女性健康照護", "IAB8": "美食與飲品", "IAB8-1": "美國料理", "IAB8-2": "燒烤料理",
    "IAB8-3": "辛香料風味料理", "IAB8-4": "中式料理", "IAB8-5": "微醺飲品", "IAB8-6": "咖啡生活",
    "IAB8-7": "異國美食探索", "IAB8-8": "手作甜點", "IAB8-9": "美食推薦", "IAB8-10": "食物過敏族群",
    "IAB8-11": "法式料理", "IAB8-12": "輕盈飲食", "IAB8-13": "義式料理", "IAB8-14": "日式料理",
    "IAB8-15": "墨西哥料理", "IAB8-16": "植物性飲食", "IAB8-17": "蔬食者", "IAB8-18": "葡萄酒",
    "IAB9": "休閒興趣嗜好", "IAB9-1": "創意科技", "IAB9-2": "手作藝術", "IAB9-3": "DIY手作串珠",
    "IAB9-4": "觀鳥", "IAB9-5": "桌遊/益智遊戲", "IAB9-6": "香氛蠟燭", "IAB9-7": "紙牌遊戲",
    "IAB9-8": "國際象棋", "IAB9-9": "雪茄", "IAB9-10": "古董收藏", "IAB9-11": "漫畫",
    "IAB9-12": "繪畫素描", "IAB9-13": "自由職業資源", "IAB9-14": "家族歷史", "IAB9-15": "出版書籍",
    "IAB9-16": "吉他", "IAB9-17": "居家音樂製作", "IAB9-18": "創新發明", "IAB9-19": "珠寶設計",
    "IAB9-20": "魔術表演", "IAB9-21": "刺繡手作", "IAB9-22": "繪畫", "IAB9-23": "攝影",
    "IAB9-24": "廣播電台", "IAB9-25": "RPG角色扮演桌遊", "IAB9-26": "科幻迷", "IAB9-27": "手帳拼貼",
    "IAB9-28": "影視編劇", "IAB9-29": "郵票收藏", "IAB9-30": "視訊和電腦遊戲", "IAB9-31": "木藝創作",
    "IAB10": "家居與園藝", "IAB10-1": "居家娛樂", "IAB10-2": "娛樂設備", "IAB10-3": "環境安全",
    "IAB10-4": "園藝", "IAB10-5": "居家維修", "IAB10-6": "家庭劇院", "IAB10-7": "室內設計&佈置",
    "IAB10-8": "景觀美化", "IAB10-9": "房屋改造", "IAB11": "法律政治議題", "IAB11-1": "移民諮詢",
    "IAB11-2": "法律議題", "IAB11-3": "政府服務", "IAB11-4": "政治", "IAB11-5": "新聞評論",
    "IAB12": "新聞與時事", "IAB12-1": "國際新聞", "IAB12-2": "國內新聞", "IAB12-3": "地方新聞",
    "IAB13": "個人財務", "IAB13-1": "新手理財", "IAB13-2": "信用卡&信貸", "IAB13-3": "財務新聞",
    "IAB13-4": "財富管理", "IAB13-5": "基金投資", "IAB13-6": "保險", "IAB13-7": "投資",
    "IAB13-8": "共同基金", "IAB13-9": "期權交易", "IAB13-10": "退休理財", "IAB13-11": "股票投資",
    "IAB13-12": "節稅規劃", "IAB14": "社會公共議題", "IAB14-1": "交友約會", "IAB14-2": "婚姻諮詢",
    "IAB14-3": "多元生活", "IAB14-4": "結婚生活", "IAB14-5": "樂齡生活", "IAB14-6": "年輕族群",
    "IAB14-7": "婚禮", "IAB14-8": "多元文化", "IAB15": "科學新知", "IAB15-1": "星座運勢",
    "IAB15-2": "生命科學", "IAB15-3": "化學", "IAB15-4": "地質學", "IAB15-5": "靈異探索",
    "IAB15-6": "物理學", "IAB15-7": "宇宙科學", "IAB15-8": "世界地理", "IAB15-9": "園藝植物",
    "IAB15-10": "天氣資訊", "IAB16": "寵物照護", "IAB16-1": "水族館", "IAB16-2": "寵物鳥",
    "IAB16-3": "萌寵毛孩-貓", "IAB16-4": "萌寵毛孩-狗", "IAB16-5": "特殊寵物", "IAB16-6": "爬蟲類",
    "IAB16-7": "寵物保健", "IAB17": "運動賽事", "IAB17-1": "賽車運動", "IAB17-2": "棒球",
    "IAB17-3": "單車運動", "IAB17-4": "體態雕塑", "IAB17-5": "拳擊", "IAB17-6": "水上運動",
    "IAB17-7": "競技啦啦隊", "IAB17-8": "攀岩", "IAB17-9": "板球", "IAB17-10": "藝術滑冰",
    "IAB17-11": "飛蠅釣", "IAB17-12": "美式足球", "IAB17-13": "淡水釣魚", "IAB17-14": "戶外釣魚",
    "IAB17-15": "高爾夫", "IAB17-16": "馬術競賽", "IAB17-17": "冰上曲棍球", "IAB17-18": "狩獵/射擊",
    "IAB17-19": "溜冰運動", "IAB17-20": "武術", "IAB17-21": "越野單車", "IAB17-22": "美式賽車",
    "IAB17-23": "國際運動盛事", "IAB17-24": "漆彈遊戲", "IAB17-25": "越野運動", "IAB17-26": "職業籃球",
    "IAB17-27": "冰上競技", "IAB17-28": "牛仔競技", "IAB17-29": "橄欖球", "IAB17-30": "慢跑日常",
    "IAB17-31": "海上休閒", "IAB17-32": "海釣", "IAB17-33": "水肺潛水", "IAB17-34": "極限滑板",
    "IAB17-35": "滑雪", "IAB17-36": "單板滑雪", "IAB17-37": "衝浪運動", "IAB17-38": "游泳運動",
    "IAB17-39": "乒乓球", "IAB17-40": "網球", "IAB17-41": "排球", "IAB17-42": "路跑",
    "IAB17-43": "水上活動", "IAB17-44": "國際足球", "IAB18": "風格與時尚", "IAB18-1": "美妝保養",
    "IAB18-2": "刺青彩繪", "IAB18-3": "風格時尚", "IAB18-4": "珠寶首飾", "IAB18-5": "流行服飾",
    "IAB18-6": "配件飾品", "IAB19": "科技趨勢", "IAB19-1": "3D繪圖", "IAB19-2": "動畫",
    "IAB19-3": "防毒軟體", "IAB19-4": "程式語言學習", "IAB19-5": "攝影器材", "IAB19-6": "智慧型手機",
    "IAB19-7": "資訊技能認證", "IAB19-8": "網路技術", "IAB19-9": "電腦配件", "IAB19-10": "科技評測",
    "IAB19-11": "資料中心", "IAB19-12": "數據管理", "IAB19-13": "數位出版", "IAB19-14": "影片編輯",
    "IAB19-15": "電郵服務", "IAB19-16": "繪圖軟體", "IAB19-17": "家庭影音娛樂", "IAB19-18": "線上協作工具",
    "IAB19-19": "Java", "IAB19-20": "JavaScript", "IAB19-21": "MacOS系統", "IAB19-22": "MP3/MIDI",
    "IAB19-23": "網路會議", "IAB19-24": "網路入門教學", "IAB19-25": "資安防護", "IAB19-26": "行動裝置",
    "IAB19-27": "PC硬體", "IAB19-28": "PHP", "IAB19-29": "行動娛樂", "IAB19-30": "共享軟體/免費軟體",
    "IAB19-31": "Unix", "IAB19-32": "VB程式設計", "IAB19-33": "數位素材", "IAB19-34": "網站設計",
    "IAB19-35": "網路搜尋", "IAB19-36": "Windows作業系統", "IAB20": "旅遊探索", "IAB20-1": "戶外探險",
    "IAB20-2": "非洲", "IAB20-3": "航空服務", "IAB20-4": "澳紐旅遊", "IAB20-5": "民宿",
    "IAB20-6": "平價旅行", "IAB20-7": "商務差旅", "IAB20-8": "美國旅遊", "IAB20-9": "露營",
    "IAB20-10": "加拿大", "IAB20-11": "加勒比海", "IAB20-12": "郵輪旅遊", "IAB20-13": "東歐",
    "IAB20-14": "歐洲", "IAB20-15": "法國", "IAB20-16": "希臘", "IAB20-17": "蜜月旅行",
    "IAB20-18": "飯店", "IAB20-19": "義大利", "IAB20-20": "日本", "IAB20-21": "拉丁美洲",
    "IAB20-22": "國家公園", "IAB20-23": "南美", "IAB20-24": "SPA水療", "IAB20-25": "主題公園",
    "IAB20-26": "家庭旅行", "IAB20-27": "英國", "IAB21": "房地產", "IAB21-1": "公寓",
    "IAB21-2": "建築師", "IAB21-3": "房屋買賣", "IAB22": "購物", "IAB22-1": "抽獎贈品",
    "IAB22-2": "折扣優惠", "IAB22-3": "優惠活動", "IAB22-4": "搜尋比價", "IAB23": "宗教與心靈成長",
    "IAB23-1": "多元信仰", "IAB23-2": "哲學思考", "IAB23-3": "佛教", "IAB23-4": "天主教",
    "IAB23-5": "基督教", "IAB23-6": "印度教", "IAB23-7": "伊斯蘭教", "IAB23-8": "猶太教",
    "IAB23-9": "摩門教", "IAB23-10": "自然信仰", "IAB24": "其他未分類內容", "IAB25": "論壇與留言",
    "IAB25-1": "未分類內容", "IAB25-2": "暴力內容防護", "IAB25-3": "成人內容過濾", "IAB25-4": "不當言論管理",
    "IAB25-5": "仇恨言論防制", "IAB25-6": "無定義內容", "IAB25-7": "獎勵機制內容", "IAB26": "限制敏感內容",
    "IAB26-1": "非法內容防護", "IAB26-2": "盜版軟體防制", "IAB26-3": "惡意軟體", "IAB26-4": "智財權保護",
    "others": "其他"
}

# Helper to reverse map
def _get_key_by_value(d: Dict[Any, str], val: str) -> Optional[Any]:
    val_clean = val.strip().lower() # Case insensitive comparison base
    for k, v in d.items():
        if v.lower() == val_clean:
            return k
    return None

class UploadParsingError(RuntimeError):
    pass

def parse_excel_df(file_bytes: bytes) -> pd.DataFrame:
    try:
        df = pd.read_excel(BytesIO(file_bytes), engine="openpyxl")
    except Exception as exc:  # pragma: no cover - defensive logging
        raise UploadParsingError("Unable to parse uploaded Excel file") from exc

    if df.empty:
        raise UploadParsingError("Sheet contains zero rows.")
    df = df.dropna(how="all")

    return df

def _validate_datetime_format(value: str, excel_row_num: int, field: str) -> str:
    if not value:
        return value

    parsed_dt = None

    # 若 Excel 讀入是 timestamp / datetime 物件
    if isinstance(value, (datetime, pd.Timestamp)):
        parsed_dt = value
    else:
        s = str(value).strip()
        # Normalization for full-width characters (e.g. ２０２４－０１－０１)
        s = unicodedata.normalize('NFKC', s)
        
        # Try multiple formats (DATE ONLY)
        formats = [
            "%Y-%m-%d",      # Date only: 2024-01-01
            "%Y/%m/%d",      # Date only slash
            "%Y.%m.%d"       # Date only dot
        ]
        
        for fmt in formats:
            try:
                parsed_dt = datetime.strptime(s, fmt)
                break
            except ValueError:
                continue

    # 成功解析後
    if parsed_dt:
        # Check date range: +/- 180 days
        now = datetime.now()
        future_limit = now + timedelta(days=180)
        past_limit = now - timedelta(days=180)
        
        if parsed_dt > future_limit:
            raise UploadParsingError(
                f"Row {excel_row_num}: 欄位「{field}」日期不能超過半年後 ({future_limit.strftime('%Y-%m-%d')})。"
            )
            
        if parsed_dt < past_limit:
            raise UploadParsingError(
                f"Row {excel_row_num}: 欄位「{field}」日期不能超過半年前 ({past_limit.strftime('%Y-%m-%d')})。"
            )

        # Return Date Only String (No Timezone Shift)
        return parsed_dt.strftime("%Y-%m-%d")

    # 都不符合 → 拋錯
    raise UploadParsingError(
        f"Row {excel_row_num}: 欄位「{field}」格式錯誤，必須為 yyyy-mm-dd hh，例如：2025-12-10 08"
    )

def excel_to_campaign_json(df: pd.DataFrame, audience_name_map: Optional[Dict[str, int]] = None) -> Dict[str, object]:
    """
    Convert the uploaded Excel sheet into the nested campaign JSON structure
    expected by Broadciel Ads v2 bulk campaign API.

    The mapping below is based on the current template header row:
      廣告活動名稱, 廣告活動ID, 每日預算(NT$), 產品類型, APP名稱, 操作系統, 品牌名稱, 廣告群組名稱,
      廣告群組ID, 網站推廣連結, 第三方點擊追蹤連結(Grouped), 第三方曝光追蹤連結(Grouped),
      行銷目標, 計費模式, 固定出價, 每日預算, 深度轉換目標, 轉換價值, 轉化目標, 開始日期, 結束日期,
      投放星期數, 投放時間段, 地理位置, 國家, 設備類型, 流量類型, 操作系統, 最高系統版本, 瀏覽器,
      年齡, 性別, 投放興趣選項, 投放興趣受眾, 自定義受眾（包含）, 自定義受眾（不包含）,
      AI語意擴充選項, AI語意擴充關鍵字, 廣告文案名稱, 廣告文案ID, 廣告類型, 廣告標題,
      廣告內文, Call to Action, 廣告素材ID

    NOTE:
      - Only clearly mappable fields are populated; many enum / type codes are left
        for the API layer to fill or default.
      - This function is intentionally defensive: missing / unexpected columns are
        tolerated and simply omitted from the output.
    """

    from urllib.parse import urlparse

    def _get_str(row: pd.Series, col: str) -> str:
        if col not in row or pd.isna(row[col]):
            return ""
        return str(row[col]).strip()

    def _get_optional_str(row: pd.Series, col: str) -> Optional[str]:
        if col not in row:
            return None
        s = _get_str(row, col)
        return s or None

    def _to_float(val: Any) -> Optional[float]:
        if pd.isna(val):
            return None
        try:
            if isinstance(val, str):
                val = val.replace(",", "")
            return float(val)
        except Exception:
            return None

    def _split_list(val: Any) -> List[str]:
        if pd.isna(val):
            return []
        if isinstance(val, (list, tuple)):
            return [str(v).strip() for v in val if str(v).strip()]
        s = str(val).strip()
        if not s:
            return []
        for sep in ["\n", ";", "；", "|"]:
            s = s.replace(sep, ",")
        parts = [p.strip() for p in s.split(",")]
        return [p for p in parts if p]

    def _parse_int_list(val: Any) -> List[int]:
        items: List[int] = []
        for token in _split_list(val):
            try:
                items.append(int(token))
            except Exception:
                # ignore non-int tokens
                continue
        return items

    def _parse_mapped_list(val: Any, mapping: Dict[Any, str]) -> List[Any]:
        """
        Parse a list of strings and map them to IDs/Codes using the provided mapping.
        Supports both Integer IDs (e.g. Platform) and String Codes (e.g. Category).
        """
        items = []
        
        # Detect Key Type from mapping (int or str)
        is_int_key = True
        if mapping:
            first_k = next(iter(mapping))
            if isinstance(first_k, str):
                is_int_key = False
                
        for token in _split_list(val):
            # 1. Reverse lookup: Chinese -> ID/Code
            mapped_id = _get_key_by_value(mapping, token)
            if mapped_id is not None:
                items.append(mapped_id)
            else:
                # 2. Fallback: Parse directly
                if is_int_key:
                    try:
                        items.append(int(token))
                    except Exception:
                        continue
                else:
                    # String code (e.g. IAB1), accept as is logic
                    # Ideally we check if token is in keys?
                    # For now just append raw string
                    items.append(token)
        return items


    # Identify columns
    # App OS: "操作系統"
    # Target OS: "受眾操作系統"
    app_os_col: Optional[str] = "操作系統"
    target_os_col: Optional[str] = "受眾操作系統"
    
    # Check existence
    if app_os_col not in df.columns:
        app_os_col = None
    if target_os_col not in df.columns:
        target_os_col = None

    campaigns: Dict[str, Dict[str, Any]] = {}

    for idx, row in df.iterrows():
        excel_row_num = idx + 2  # header assumed row 1

        campaign_name = _get_str(row, "廣告活動名稱")
        if not campaign_name:
            # skip empty campaign rows
            continue

        product_type_raw = _get_str(row, "產品類型")
        product_type = product_type_raw.lower()
        app_name_raw = _get_optional_str(row, "APP名稱")
        app_os_raw = _get_optional_str(row, app_os_col) if app_os_col else None

        app_present = bool(app_name_raw)
        os_present = bool(app_os_raw)

        # 保留原本邏輯：web 活動不能填 APP / OS
        if product_type == "web" and (app_present or os_present):
            raise UploadParsingError(
                f"Row {excel_row_num}: Web campaign must not have APP名稱 or 操作系統 filled."
            )

        # 初始化或取得 campaign 物件
        if campaign_name not in campaigns:
            campaign: Dict[str, Any] = {
                # 廣告活動名稱 → cpg_name
                "cpg_name": campaign_name,
            }

            # 廣告活動ID → cpg_id
            cpg_id_val = _get_optional_str(row, "廣告活動ID")
            if cpg_id_val is not None:
                try:
                    campaign["cpg_id"] = int(float(cpg_id_val))
                except Exception:
                    pass

            # 活動層級預算
            cpg_budget_val = row.get("每日預算(NT$)")
            cpg_budget = _to_float(cpg_budget_val)
            if cpg_budget is not None:
                # 每日預算(NT$) → day_budget
                campaign["day_budget"] = cpg_budget

            # 品牌名稱 → sponsored
            brand = _get_optional_str(row, "品牌名稱")
            if brand:
                campaign["sponsored"] = brand

            # 主網域名稱 → adomain
            adomain = _get_optional_str(row, "主網域名稱")
            if adomain is not None:
                # 移除 http:// 或 https://
                cleaned_adomain = adomain.lower()
                for prefix in ["https://", "http://"]:
                    if cleaned_adomain.startswith(prefix):
                        cleaned_adomain = cleaned_adomain[len(prefix):]
                # 也要移除結尾的 / (如果有的話)
                cleaned_adomain = cleaned_adomain.rstrip("/")
                
                campaign["adomain"] = cleaned_adomain
            else:
                raise UploadParsingError(
                        f"Row {excel_row_num}: 主網域名稱必須填寫。"
                    )

            # 產品類型 → ad_channel (1=app, 2=web)
            if product_type == "app":
                campaign["ad_channel"] = 1
            elif product_type == "web":
                campaign["ad_channel"] = 2

            # app-only fields (APP名稱 / 操作系統)
            if product_type == "app":
                app_obj: Dict[str, Any] = {}
                if app_name_raw:
                    # APP名稱 → app.ad_target
                    app_obj["ad_target"] = app_name_raw

                # 操作系統 → app.ad_platform (1=iOS, 2=Android)
                if app_os_raw:
                    os_name = app_os_raw.lower()
                    if "ios" in os_name:
                        app_obj["ad_platform"] = 1
                    elif "android" in os_name:
                        app_obj["ad_platform"] = 2

                # 規則：ad_channel 為 1 (app) 時，ad_target / ad_platform 不可為空
                if not app_obj.get("ad_target") or not app_obj.get("ad_platform"):
                    raise UploadParsingError(
                        f"Row {excel_row_num}: APP 活動必須填寫 APP名稱 與 操作系統。"
                    )

                campaign["app"] = app_obj

            # 廣告活動狀態 -> cpg_status (1=開啟, 2=關閉)
            cpg_status_raw = _get_optional_str(row, "廣告活動狀態")
            if cpg_status_raw:
                 # Check for text or numeric representation
                 v = cpg_status_raw.strip()
                 if v == "開啟":
                      campaign["cpg_status"] = 1
                 elif v == "關閉":
                      campaign["cpg_status"] = 2
                 elif v in ("1", "1.0"):
                      campaign["cpg_status"] = 1
                 elif v in ("2", "2.0"):
                      campaign["cpg_status"] = 2
            else:
                 # Default to 1 (Active) ONLY if not specified
                 # Wait, logic in CampaignBulkProcessor defaults to 1 if missing.
                 # But we need to ensure "關閉" becomes 2 here.
                 # If empty, we leave it out and let processor decide? 
                 # Or set default here?
                 # Requirement: "新增時預設開啟，但填關閉必須是關閉"
                 # So default 1 here is fine, but if "關閉" is present it MUST be 2.
                 
                 # CampaignBulkProcessor logic:
                 # if "cpg_status" in campaign_data: set it
                 # else: set 1
                 
                 # So if we don't set it here, it defaults to 1.
                 # If we set 2 here, it stays 2.
                 # This logic is correct for the requirement.
                 pass
            
            campaigns[campaign_name] = campaign
        else:
            campaign = campaigns[campaign_name]

        #
        # 廣告群組 (ad_group) 層級
        #
        group_name = _get_str(row, "廣告群組名稱") or "Default Group"
        ad_groups: List[Dict[str, Any]] = campaign.setdefault("ad_group", [])

        # 找到 / 建立 group
        group: Optional[Dict[str, Any]] = None
        for g in ad_groups:
            if g.get("group_name") == group_name:
                group = g
                break
        if group is None:
            group = {
                # 廣告群組名稱 → group_name
                "group_name": group_name
            }
            # 廣告群組ID → group_id
            group_id_val = _get_optional_str(row, "廣告群組ID")
            if group_id_val is not None:
                try:
                    group["group_id"] = int(float(group_id_val))
                except Exception:
                    pass
            
            # 廣告群組狀態 -> group_status (1=開啟, 2=關閉)
            group_status_raw = _get_optional_str(row, "廣告群組狀態")
            if group_status_raw:
                 v = group_status_raw.strip()
                 if v == "開啟":
                      group["group_status"] = 1
                 elif v == "關閉":
                      group["group_status"] = 2
                 elif v in ("1", "1.0"):
                      group["group_status"] = 1
                 elif v in ("2", "2.0"):
                      group["group_status"] = 2

            ad_groups.append(group)

        #
        # URL / tracking
        #
        # 網站推廣連結 → target_info (主要推廣頁面)
        target_info_url = _get_optional_str(row, "網站推廣連結")
        if target_info_url:
            group["target_info"] = target_info_url

        # 第三方點擊追蹤連結(Grouped) → click_url, only https
        click_urls = [u.strip() for u in _split_list(row.get("第三方點擊追蹤連結(Grouped)")) if u and str(u).strip().startswith("https://")]
        if click_urls:
            group["click_url"] = click_urls

        # 第三方曝光追蹤連結(Grouped) → impression_url, only https and formatted as <img src="url">
        imp_urls = [u.strip() for u in _split_list(row.get("第三方曝光追蹤連結(Grouped)")) if u and str(u).strip().startswith("https://")]
        if imp_urls:
            group["impression_url"] = [{"type": 2, "value": f'<img src="{u}">'} for u in imp_urls]

        #
        # Budget block
        #
        budget: Dict[str, Any] = group.setdefault("budget", {})

        # 行銷目標 → market_goal
        marketing_goal = _get_optional_str(row, "行銷目標")
        if marketing_goal:
            market_target_map = {
                "品牌知名度": 1,
                "電商網上購買": 2,
                "增加網站流量": 3,
                "開發潛在客戶": 5,
                "網站互動": 6,
            }
            market_target = market_target_map.get(marketing_goal)
            if market_target is None:
                raise UploadParsingError(
                    f"Row {excel_row_num}: 不支援的行銷目標「{marketing_goal}」，請確認是否拼寫正確。"
                )
            budget["market_target"] = market_target

        # 計費模式 → rev_type
        billing_type = _get_optional_str(row, "計費模式")
        if billing_type:
            type_map = {"CPM": 2, "CPC": 3}
            budget["rev_type"] = type_map.get(billing_type, billing_type)

        price_val = row.get("固定出價")
        price = _to_float(price_val)
        if price is not None:
            # 固定出價 → price
            budget["price"] = price

        group_day_budget_val = row.get("每日預算")
        group_day_budget = _to_float(group_day_budget_val)
        if group_day_budget is not None:
            # 廣告群組「每日預算」 → budget.day_budget
            budget["day_budget"] = group_day_budget

        # 深度轉換目標 / 轉換價值 / 轉化目標 → conversion_goal 物件
        depth_goal = _get_optional_str(row, "深度轉換目標")  # → conversion_goal
        conv_value = _to_float(row.get("轉換價值"))  # → target_value
        conv_goal = _get_optional_str(row, "轉化目標")  # → convert_event

        conversion_goal: Dict[str, Any] = {}
        if depth_goal is not None:
            #conversion_goal["conversion_goal"] = depth_goal
            goal_type_map = {
                "帳戶預設設定": 0,
                "所有轉換": 1,
                "指定轉換目標": 2,
            }
            t = goal_type_map.get(depth_goal)
            if t is None:
                raise UploadParsingError(
                    f"Row {excel_row_num}: 不支援的深度轉換目標「{depth_goal}」，請確認是否拼寫正確。"
                )
            conversion_goal["type"] = t
            if t == 1 and conv_value is None:
                raise UploadParsingError(
                    f"Row {excel_row_num}: 深度轉換目標為「所有轉換」時，必須填寫『轉換價值』。"
                )
            if t == 2:
                missing = []
                if conv_value is None:
                    missing.append("轉換價值")
                if conv_goal is None:
                    missing.append("轉化目標")
                if missing:
                    raise UploadParsingError(
                        f"Row {excel_row_num}: 深度轉換目標為「指定轉換目標」時，必須填寫「{'、'.join(missing)}」。"
                    )

        # target_value：只有在 type != 0 時才輸出
        if conv_value is not None and t != 0:
            conversion_goal["target_value"] = conv_value

        # convert_event：只有在 type != 0 時才輸出
        if conv_goal is not None and t != 0:
            conv_goal_map = {
                "點擊數": 11,
                "網頁瀏覽": 13,
                "完成註冊": 6,
                "搜尋": 5,
                "收藏": 3,
                "加入購物車": 4,
                "開始結帳": 2,
                "完成結帳": 1,
            }
            cg = conv_goal_map.get(conv_goal)
            if cg is None:
                raise UploadParsingError(
                    f"Row {excel_row_num}: 不支援的轉換目標「{conv_goal}」，請確認是否拼寫正確。"
                )
            conversion_goal["convert_event"] = cg

        if conversion_goal:
            budget["conversion_goal"] = conversion_goal

        #
        # Schedule
        #
        schedule: Dict[str, Any] = group.setdefault("schedule", {})

        # 開始日期 / 結束日期
        start_date_raw = row.get("開始日期")
        end_date_raw = row.get("結束日期")

        if not start_date_raw or pd.isna(start_date_raw) or str(start_date_raw).strip() == "":
            raise UploadParsingError(f"Row {excel_row_num}: 「開始日期」為必填欄位。")
        else:
            start_date = _validate_datetime_format(str(start_date_raw), excel_row_num, "開始日期")
            schedule["start_date"] = start_date

        if not end_date_raw or pd.isna(end_date_raw) or str(end_date_raw).strip() == "":
            raise UploadParsingError(f"Row {excel_row_num}: 「結束日期」為必填欄位。")
        else:
            end_date = _validate_datetime_format(str(end_date_raw), excel_row_num, "結束日期")
            schedule["end_date"] = end_date

        # 投放星期數 → week_days
        week_days_raw = row.get("投放星期數")
        week_days = _parse_int_list(week_days_raw)
        if week_days:
            schedule["week_days"] = week_days

        # 投放時間段 → hours
        hours_raw = row.get("投放時間段")
        hours = _parse_int_list(hours_raw)
        if hours:
            # UTC+8 換成 UTC (local - 8)
            hours_utc = [(h - 8) % 24 for h in hours]
            schedule["hours"] = hours_utc

        #
        # Location
        #
        location: Dict[str, Any] = group.setdefault("location", {})
        # 地理位置 → country_type
        country_type_raw = _get_optional_str(row, "地理位置")
        if country_type_raw:
            country_type_map = {
                "包含": 1,
                "不包含": 2
            }
            country_type = country_type_map.get(country_type_raw)
            if country_type is None:
                raise UploadParsingError(
                    f"Row {excel_row_num}: 不支援的地理位置「{country_type_raw}」，請確認是否拼寫正確。"
                )
            location["country_type"] = country_type
        # 國家 → country
        country = _split_list(row.get("國家"))
        if country:
            location["country"] = country

        #
        # Audience targeting
        #
        audience: Dict[str, Any] = group.setdefault("audience_target", {})

        # 設備類型 → device_type
        device_types = _split_list(row.get("設備類型"))
        if device_types:
            device_types_int = []
            for d in device_types:
                # Try map lookup first
                mapped_id = _get_key_by_value(DEVICE_MAP, d)
                if mapped_id is not None:
                    device_types_int.append(mapped_id)
                    continue
                try:
                    device_types_int.append(int(d))
                except Exception:
                    continue
            audience["device_type"] = device_types_int

        # 流量類型 → traffic_type, force int
        traffic_types = _split_list(row.get("流量類型"))
        if traffic_types:
            traffic_types_int = []
            for t in traffic_types:
                mapped_id = _get_key_by_value(TRAFFIC_MAP, t)
                if mapped_id is not None:
                    traffic_types_int.append(mapped_id)
                    continue
                try:
                    traffic_types_int.append(int(t))
                except Exception:
                    continue
            audience["traffic_type"] = traffic_types_int

        # 受眾層級 OS / 平台: 操作系統(第二欄) → platform, force int
        target_os_raw = (
            _get_optional_str(row, target_os_col) if target_os_col else None
        )
        if target_os_raw:
            audience["platform"] = _parse_mapped_list(target_os_raw, PLATFORM_MAP)

        # 最高系統版本 → os_version as {min, max}
        max_os_ver_raw = row.get("最高系統版本")
        min_version = 0
        max_version = 0
        audience["os_version"] = {"min": min_version, "max": max_version}

        # 瀏覽器 → browser
        audience["browser"] = _parse_mapped_list(row.get("瀏覽器"), BROWSER_MAP)

        # 年齡 → age
        ages = _split_list(row.get("年齡"))
        if ages:
            ages_int = []
            for a in ages:
                mapped_id = _get_key_by_value(AGE_MAP, a)
                if mapped_id is not None:
                    ages_int.append(mapped_id)
                    continue
                try:
                    ages_int.append(int(a))
                except Exception:
                    continue
            audience["age"] = ages_int

        # 性別 → gender
        genders = _split_list(row.get("性別"))
        if genders:
            genders_int = []
            for g in genders:
                mapped_id = _get_key_by_value(GENDER_MAP, g)
                if mapped_id is not None:
                    genders_int.append(mapped_id)
                    continue
                try:
                    genders_int.append(int(g))
                except Exception:
                    continue
            audience["gender"] = genders_int

        # 興趣 / IAB / 關鍵字類
        # 投放興趣選項 → category.type (包含=1, 不包含=2)
        # 投放興趣受眾 → category.value (IAB list)
        interest_opt = _get_optional_str(row, "投放興趣選項")
        interest_audience = _split_list(row.get("投放興趣受眾"))
        
        if interest_opt or interest_audience:
            category: Dict[str, Any] = {}
            
            # Parse type from 投放興趣選項 (包含 or 不包含)
            if interest_opt:
                if interest_opt == "包含":
                    category["type"] = 1
                elif interest_opt == "不包含":
                    category["type"] = 2
                else:
                    raise UploadParsingError(
                        f"Row {excel_row_num}: 投放興趣選項必須為「包含」或「不包含」。"
                    )
            
            # Use 投放興趣受眾 (IAB) as the value list (Map Names -> Codes)
            if interest_audience:
                # Note: original list is already split by _split_list in interest_audience
                # But _parse_mapped_list expects raw value usually? 
                # Actually _parse_mapped_list calls _split_list again internally.
                # So we can pass row.get("投放興趣受眾") directly.
                category["value"] = _parse_mapped_list(row.get("投放興趣受眾"), CATEGORY_MAP)
            
            if category:
                audience["category"] = category

            if category:
                audience["category"] = category

        # APP/網站篩選 (Site)
        # 欄位: "APP/網站篩選選項" (type: 1=包含, 2=不包含), "APP/網站篩選" (url)
        site_opt_raw = row.get("APP/網站篩選選項")
        site_url_raw = row.get("APP/網站篩選")
        
        if site_url_raw and not pd.isna(site_url_raw):
            site_obj = {}
            # Parse Type
            s_type = 1 # default
            if site_opt_raw:
                s_opt_str = str(site_opt_raw).strip()
                if s_opt_str == "包含" or s_opt_str == "1":
                    s_type = 1
                elif s_opt_str == "不包含" or s_opt_str == "2":
                    s_type = 2
            
            site_obj["type"] = s_type
            site_obj["url"] = str(site_url_raw).strip()
            # audience["site"] = site_obj
            


        # AI語意擴充選項 / AI語意擴充關鍵字 → keywords: { type, value }
        # AI語意擴充選項 should be "1" or "2"
        ai_expand_opt_raw = row.get("AI語意擴充選項")
        ai_expand_keywords = _split_list(row.get("AI語意擴充關鍵字"))  # → keywords.value
        
        opt_str = ""
        if ai_expand_opt_raw is not None and not pd.isna(ai_expand_opt_raw):
            opt_str = str(ai_expand_opt_raw).strip().lower()

        # If both option and keywords are empty, skip silently
        if opt_str == "" and not ai_expand_keywords:
            pass
        else:
            kw_type: Optional[int] = None
            if opt_str == "" and ai_expand_keywords:
                # keywords provided but no type
                raise UploadParsingError(
                    f"Row {excel_row_num}: AI語意擴充選項必須為 1 或 2。"
                )
            elif opt_str:
                if opt_str in ("1", "1.0") or opt_str == "包含":
                    kw_type = 1
                elif opt_str in ("2", "2.0") or opt_str == "不包含":
                    kw_type = 2
                else:
                    raise UploadParsingError(
                        f"Row {excel_row_num}: AI語意擴充選項必須為 1 或 2。"
                    )

            # Only add keywords if type is set
            if kw_type is not None:
                keywords: Dict[str, Any] = {
                    "type": kw_type,
                    "value": ai_expand_keywords  # Always set value, even if empty array
                }
                audience["keywords"] = keywords

        # 自定義受眾（包含）→ pixel_audience with type: 1
        # 自定義受眾（不包含）→ pixel_audience with type: 2
        pixel_audience_include = _split_list(row.get("自定義受眾（包含）"))
        pixel_audience_exclude = _split_list(row.get("自定義受眾（不包含）"))
        
        pixel_audience_list: List[Dict[str, int]] = []
        
        def _resolve_audience_id(val: str) -> Optional[int]:
            # 1. Try integer
            try:
                return int(val)
            except (ValueError, TypeError):
                pass
            
            # 2. Try map look up
            if audience_name_map:
                # remove whitespace just in case
                v_clean = val.strip()
                if v_clean in audience_name_map:
                    return audience_name_map[v_clean]
            
            return None

        # Add include audiences (type: 1)
        for val_str in pixel_audience_include:
            aid = _resolve_audience_id(val_str)
            if aid is not None:
                pixel_audience_list.append({"id": aid, "type": 1})
            else:
                # Optional: log warning or raise error if name not found?
                # For now, we just skip invalid ones as per original logic, 
                # but maybe logging would be better.
                pass
        
        # Add exclude audiences (type: 2)
        for val_str in pixel_audience_exclude:
            aid = _resolve_audience_id(val_str)
            if aid is not None:
                pixel_audience_list.append({"id": aid, "type": 2})
            else:
                pass
        
        if pixel_audience_list:
            audience["pixel_audience"] = pixel_audience_list

        #
        # Ad assets (creatives) per ad group
        #
        ad_assets: List[Dict[str, Any]] = group.setdefault("ad_asset", [])

        cr_name = _get_str(row, "廣告文案名稱")
        cr_title = _get_str(row, "廣告標題")
        cr_desc = _get_str(row, "廣告內文")
        cr_btn = _get_str(row, "Call to Action")
        cr_iab = _get_str(row, "廣告類型")  # per mapping request

        # 廣告文案狀態 (Creative Status)
        cr_status_raw = _get_optional_str(row, "廣告文案狀態")
        cr_status_val = None
        if cr_status_raw:
            v_cr = cr_status_raw.strip()
            if v_cr == "開啟":
                cr_status_val = 1
            elif v_cr == "關閉":
                cr_status_val = 2
            elif v_cr == "已過期":
                cr_status_val = 4
            elif v_cr in ("1", "1.0"):
                cr_status_val = 1
            elif v_cr in ("2", "2.0"):
                cr_status_val = 2
            elif v_cr in ("4", "4.0"):
                cr_status_val = 4

        cr_mt_raw = row.get("廣告素材ID")
        cr_mt_id: Optional[int | str] = None
        if cr_mt_raw is not None and not pd.isna(cr_mt_raw):
            try:
                # Handle "101.0" or "101"
                cr_mt_id = int(float(str(cr_mt_raw).strip()))
            except Exception:
                # If parsing fails, do NOT fallback to string.
                # Valid ID must be int.
                cr_mt_id = None

        # Determine Creative ID (cr_id) for Update
        cr_id_raw = row.get("廣告文案ID")
        cr_id: Optional[int] = None
        if cr_id_raw is not None and not pd.isna(cr_id_raw):
             try:
                 # Handle "123.0"
                 cr_id = int(float(str(cr_id_raw).strip()))
             except:
                 pass

        # Add asset only if at least one meaningful field is present
        # Include cr_id in check
        if any([cr_name, cr_title, cr_desc, cr_btn, cr_iab, cr_mt_id, cr_id]):
            asset: Dict[str, Any] = {
                #"group_id": group.get("group_id"),
                "cr_name": cr_name,
                "cr_title": cr_title,
                "cr_desc": cr_desc,
                "cr_btn_text": cr_btn,
                "iab": cr_iab,
                "cr_mt_id": cr_mt_id if cr_mt_id is not None else 0,
                "cr_icon_id": 0,
            }
            if cr_id:
                asset["cr_id"] = cr_id
            if cr_status_val is not None:
                # Use ad_status generally, or whatever the API needs. 
                # Broadciel Client usually uses cr_status or ad_status depending on endpoint.
                # CampaignBulkProcessor will extract it.
                asset["cr_status"] = cr_status_val
                
            ad_assets.append(asset)

    return {"campaign": list(campaigns.values())}


def dataframe_preview(df: pd.DataFrame, limit: int = 50) -> Dict[str, object]:
    rows = df.fillna("").astype(str).head(limit).to_dict(orient="records")
    return {
        "columns": df.columns.tolist(),
        "rows": rows,
        "total_rows": len(df.index),
        "preview_count": len(rows),
    }


def parse_excel(file_bytes: bytes) -> Dict[str, object]:
    try:
        df = pd.read_excel(BytesIO(file_bytes), engine="openpyxl")
    except Exception as exc:  # pragma: no cover - defensive logging
        raise UploadParsingError("Unable to parse uploaded Excel file") from exc

    if df.empty:
        raise UploadParsingError("Sheet contains zero rows.")

    return dataframe_preview(df)


def generate_excel_from_api_data(
    campaigns: List[Dict[str, Any]],
    ad_groups: List[Dict[str, Any]],
    ad_creatives: List[Dict[str, Any]],
    audience_id_map: Optional[Dict[int, str]] = None
) -> bytes:
    """
    Generate an Excel file (bytes) from the API data structures, 
    matching the upload format exactly (46+3 columns) with Data Validation.
    """
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # helper for status check
    def _is_archived(obj, status_key):
        # 3 = Archived, skip
        val = obj.get(status_key)
        return val == 3

    # 1. Indexing for fast lookup
    # Normalize IDs to int to avoid hash mismatches (e.g. "123" vs 123)
    cpg_map = {}
    for c in campaigns:
        try:
             cid = int(c.get("cpg_id"))
             cpg_map[cid] = c
        except (ValueError, TypeError):
             continue
    
    layout: Dict[int, Any] = {}
    
    for cpg_id, c in cpg_map.items():
        if _is_archived(c, "cpg_status"):
            continue
        layout[cpg_id] = {"self": c, "groups": {}}
        
    for g in ad_groups:
        if _is_archived(g, "group_status"):
            continue
            
        try:
            cpg_id = int(g.get("cpg_id"))
            grp_id = int(g.get("group_id"))
        except (ValueError, TypeError):
            continue
            
        if cpg_id in layout:
            layout[cpg_id]["groups"][grp_id] = {"self": g, "creatives": []}
        else:
            pass
            
    # Track seen creatives to avoid duplicates (e.g. if API returns dupes)
    seen_creatives = set()

    for cr in ad_creatives:
        # Check both ad_status and cr_status, sometimes one is used
        status = cr.get("cr_status") or cr.get("ad_status")
        if status == 3:
            continue
        
        try:
            cpg_id = int(cr.get("cpg_id"))
            grp_id = int(cr.get("group_id"))
            cr_id = int(cr.get("cr_id")) if cr.get("cr_id") is not None else None
        except (ValueError, TypeError):
             continue
        
        # Deduplication check
        if cr_id and (grp_id, cr_id) in seen_creatives:
            continue
            
        if cpg_id in layout:
            if grp_id in layout[cpg_id]["groups"]:
                layout[cpg_id]["groups"][grp_id]["creatives"].append(cr)
                if cr_id:
                    seen_creatives.add((grp_id, cr_id))
                
            else:
                 pass
        else:
             pass
             
    # 2. Flatten to Rows
    rows = []
    
    # Track previous IDs for Sparse Writing (Smart Edit)
    last_cpg_id = None
    last_grp_id = None
    
    # Define Column Mappings (same as before)
    def _map_product_type(val):
        return "app" if val == 1 else "web" 
        
    def _map_os(val):
        if val == 1: return "iOS"
        if val == 2: return "Android"
        if val == 3: return "Others"
        return "" 

    def _map_status(val):
        if val == 1: return "開啟"
        if val == 2: return "關閉"
        if val == 4: return "已過期"
        return "" 

    def _map_market_goal(val):
        m = {1: "品牌知名度", 2: "電商網上購買", 3: "增加網站流量", 5: "開發潛在客戶", 6: "網站互動"}
        return m.get(val, "")

    def _map_billing(val):
        if val == 2: return "CPM"
        if val == 3: return "CPC"
        return ""
        
    def _map_conversion_goal_type(val):
        m = {0: "帳戶預設設定", 1: "所有轉換", 2: "指定轉換目標"}
        return m.get(val, "")
        
    def _map_convert_event(val):
        m = {11: "點擊數", 13: "網頁瀏覽", 6: "完成註冊", 5: "搜尋", 3: "收藏", 4: "加入購物車", 2: "開始結帳", 1: "完成結帳"}
        return m.get(val, "")

    def _map_country_type(val):
        if val == 1: return "包含"
        if val == 2: return "不包含"
        return ""
        
    def _map_list(lst, mapping: Optional[Dict[Any, str]] = None):
        if not lst: return ""
        if mapping:
            # Check key type
            is_int_key = True
            first_k = next(iter(mapping)) if mapping else None
            if isinstance(first_k, str):
                is_int_key = False
            
            mapped = []
            for x in lst:
                if is_int_key:
                    try:
                        k = int(x)
                        mapped.append(mapping.get(k, str(x)))
                    except:
                        mapped.append(str(x))
                else:
                    # String key
                    k = str(x)
                    mapped.append(mapping.get(k, k))
            return ",".join(mapped)
        
        return ",".join(str(x) for x in lst)
        
    def _map_hours(utc_hours):
        if not utc_hours: return ""
        # User requested NO timezone adjustment (keep UTC)
        local_hours = [str(h) for h in utc_hours]
        return ",".join(local_hours)
        
    def _map_ai_opt(val):
        if not val: return ""
        if val == 1: return "包含"
        if val == 2: return "不包含"
        if val == "1": return "包含"
        if val == "2": return "不包含"
        return str(val)

    def _map_list_obj_value(lst):
        urls = []
        for x in lst:
            v = x.get("value", "")
            if v.startswith('<img src="') and v.endswith('">'):
                v = v[10:-2]
            elif v.startswith("<img src='") and v.endswith("'>"):
                v = v[10:-2]
            urls.append(v)
        return ",".join(urls)


    # Re-implementing the nested loop structure with correct sparse check placement
    
    for cpg_id, c_node in layout.items():
        c = c_node["self"]
        groups_node = c_node["groups"]
        
        # Prepare Campaign Values
        c_name = c.get("cpg_name", "")
        c_id = c.get("cpg_id", "")
        c_status = _map_status(c.get("cpg_status"))
        c_budget = c.get("day_budget", "")
        c_domain = c.get("adomain", "")
        c_prod_type = _map_product_type(c.get("ad_channel"))
        app_info = c.get("app", {})
        c_app_name = app_info.get("ad_target", "")
        c_os = _map_os(app_info.get("ad_platform"))
        c_brand = c.get("sponsored", "")
        
        c_cols_data = [
            c_name, c_id, c_status,
            c_budget, c_domain, c_prod_type, c_app_name, c_os, c_brand
        ]
        
        if not groups_node:
            # No groups, write campaign data only
            row = c_cols_data + [""] * 32 + [""] * 8
            rows.append(row)
            continue
            
        for grp_id, g_node in groups_node.items():
            g = g_node["self"]
            creatives = g_node["creatives"]
            
            # Prepare Group Values
            g_name = g.get("group_name", "")
            g_id = g.get("group_id", "")
            g_status = _map_status(g.get("group_status"))
            g_target = g.get("target_info", "")
            g_click = ",".join(g.get("click_url", []))
            g_imp = _map_list_obj_value(g.get("impression_url", []))
            b_obj = g.get("budget", {})
            g_market = _map_market_goal(b_obj.get("market_target"))
            g_rev = _map_billing(b_obj.get("rev_type"))
            g_price = b_obj.get("price", "")
            g_day_budget = b_obj.get("day_budget", "")
            cv_obj = b_obj.get("conversion_goal", {})
            g_depth = _map_conversion_goal_type(cv_obj.get("type"))
            g_cv_val = cv_obj.get("target_value", "")
            g_cv_event = _map_convert_event(cv_obj.get("convert_event"))
            sched = g.get("schedule", {})
            g_start = sched.get("start_date", "")
            g_end = sched.get("end_date", "")
            g_week = _map_list(sched.get("week_days"))
            g_hours = _map_hours(sched.get("hours"))
            loc = g.get("location", {})
            g_loc_type = _map_country_type(loc.get("country_type"))
            g_country = _map_list(loc.get("country"))
            aud = g.get("audience_target", {})
            a_device = _map_list(aud.get("device_type"), DEVICE_MAP)
            a_traffic = _map_list(aud.get("traffic_type"), TRAFFIC_MAP)
            a_platform = _map_list(aud.get("platform"), PLATFORM_MAP)
            a_os_ver = "" 
            a_browser = _map_list(aud.get("browser"), BROWSER_MAP)
            a_age = _map_list(aud.get("age"), AGE_MAP)
            a_gender = _map_list(aud.get("gender"), GENDER_MAP)
            cat = aud.get("category", {})
            a_cat_type = _map_country_type(cat.get("type"))
            a_cat_val = _map_list(cat.get("value"), CATEGORY_MAP)
            if not a_cat_val:
                a_cat_type = ""
                
            # Site Filter (APP/網站篩選)
            site_obj = aud.get("site", {})
            a_site_type = _map_country_type(site_obj.get("type")) # reusing map for 1=包含, 2=不包含
            a_site_url = site_obj.get("url", "")
            if not a_site_url:
                a_site_type = ""
                
            pix_include = []
            pix_exclude = []
            for p in aud.get("pixel_audience", []):
                pid = p.get("id")
                ptype = p.get("type")
                if ptype == 1: pix_include.append(pid)
                elif ptype == 2: pix_exclude.append(pid)
            a_pix_inc = _map_list(pix_include, audience_id_map)
            a_pix_exc = _map_list(pix_exclude, audience_id_map)
            kw = aud.get("keywords", {})
            a_ai_type = _map_ai_opt(kw.get("type", ""))
            a_ai_val = _map_list(kw.get("value"))
            
            # Request: If keywords (value) is empty, Option (type) should be empty
            if not a_ai_val:
                a_ai_type = ""
            
            g_cols_data = [
                g_name, g_id, g_status,
                g_target, g_click, g_imp,
                g_market, g_rev, g_price, g_day_budget, g_depth, g_cv_val, g_cv_event, g_start, g_end,
                g_week, g_hours, g_loc_type, g_country, a_device, a_traffic, a_platform, a_os_ver, a_browser,
                a_age, a_gender, a_cat_type, a_cat_val, a_site_type, a_site_url, a_pix_inc, a_pix_exc,
                a_ai_type, a_ai_val
            ]
            
            if not creatives:
                c_cols = c_cols_data 
                g_cols = g_cols_data 
                     
                row = c_cols + g_cols + [""] * 8
                rows.append(row)
                continue
                
            for cr in creatives:
                cr_name = cr.get("cr_name", "")
                cr_id = cr.get("cr_id", "")
                cr_status = _map_status(cr.get("cr_status") or cr.get("ad_status"))
                cr_iab = cr.get("iab", "")
                cr_title = cr.get("cr_title", "")
                cr_desc = cr.get("cr_desc", "")
                cr_btn = cr.get("cr_btn_text", "")
                cr_mt = cr.get("cr_mt_id") or cr.get("cr_mt") or ""
                
                cr_cols = [
                    cr_name, cr_id, cr_status,
                    cr_iab, cr_title, cr_desc, cr_btn, cr_mt
                ]
                
                c_cols = c_cols_data
                g_cols = g_cols_data 
                    
                row = c_cols + g_cols + cr_cols
                rows.append(row)

    # 3. Create DataFrame
    # group_status inserted at index 11
    # creative_status inserted at index 43
    columns = [
        "廣告活動名稱", "廣告活動ID", "廣告活動狀態",
        "每日預算(NT$)", "主網域名稱", "產品類型", "APP名稱", "操作系統", "品牌名稱",
        "廣告群組名稱", "廣告群組ID", "廣告群組狀態", "網站推廣連結", "第三方點擊追蹤連結(Grouped)",
        "第三方曝光追蹤連結(Grouped)",
        "行銷目標", "計費模式", "固定出價", "每日預算", "深度轉換目標", "轉換價值", "轉化目標", "開始日期", "結束日期",
        "投放星期數", "投放時間段", "地理位置", "國家", "設備類型", "流量類型", "受眾操作系統", "最高系統版本", "瀏覽器",
        "年齡", "性別", "投放興趣選項", "投放興趣受眾", "APP/網站篩選選項", "APP/網站篩選", "自定義受眾（包含）", "自定義受眾（不包含）",
        "AI語意擴充選項", "AI語意擴充關鍵字", 
        "廣告文案名稱", "廣告文案ID", "廣告文案狀態",
        "廣告類型", "廣告標題", "廣告內文", "Call to Action", "廣告素材ID"
    ]
    
    df = pd.DataFrame(rows, columns=columns)
    
    # 4. Write to Bytes using OpenPyXL directly for advanced features
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # --- Styles ---
        from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
        
        # Fonts
        font_header = Font(bold=True, size=12)
        font_data = Font(size=12)
        
        # Colors (Light Pastel tones)
        # Campaign: Light Blue
        color_cpg = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        # Ad Group: Light Green
        color_grp = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        # Creative: Light Yellow/Orange
        color_crt = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # Border Styles
        # Data Gridline: Light Grey (Simulating default gridlines)
        grid_border = Border(
            left=Side(style='thin', color='D9D9D9'), 
            right=Side(style='thin', color='D9D9D9'), 
            top=Side(style='thin', color='D9D9D9'), 
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Adjust Header Row Height (1.5x approx default 20 -> 30)
        worksheet.row_dimensions[1].height = 30
        
        # Alignment for Header
        align_center = Alignment(vertical='center', horizontal='center')

        # Apply Header Colors and Font
        # Helper to apply fill and font to header range
        def _style_header(start_col_idx, end_col_idx, fill):
            # openpyxl is 1-based indexing for rows/cols
            for c_idx in range(start_col_idx, end_col_idx + 1):
                cell = worksheet.cell(row=1, column=c_idx)
                cell.fill = fill
                cell.font = font_header
                # cell.border = header_border  # Removed per request
                cell.border = grid_border # Use same light grey border as data
                cell.alignment = align_center

        _style_header(1, 9, color_cpg)   # Campaign
        _style_header(10, 41, color_grp) # Ad Group
        _style_header(42, 49, color_crt) # Creative
        
        
        # Apply Font, Gridlines AND Background Colors to Data Range
        # Iterate over all rows from 2 to max_row
        
        # Determine strict data range
        max_r = len(df) + 1 # +1 for header
        max_c = len(columns)
        
        for r_idx in range(2, max_r + 1):
            for c_idx in range(1, max_c + 1):
                cell = worksheet.cell(row=r_idx, column=c_idx)
                cell.font = font_data
                # Apply light grey gridline
                cell.border = grid_border
                
                # Apply Background Color matching Header
                if 1 <= c_idx <= 9:
                    cell.fill = color_cpg
                elif 10 <= c_idx <= 41:
                    cell.fill = color_grp
                elif 42 <= c_idx <= 49:
                    cell.fill = color_crt
                
        # --- 2. Hide Default Gridlines (Requests 2 & 3) ---
        # "完全沒資料的row/col 把預設的框線也消除，全白的那種"
        worksheet.sheet_view.showGridLines = False

        # --- 1. Auto-fit Column Width (Request 1) ---
        # Heuristic: Iterate rows to find max length per column.
        # Since we have sparse data (blanks), we check valid cells.
        # openpyxl requires setting 'worksheet.column_dimensions[letter].width'
        
        # Initialize max_lengths with header lengths
        # columns is list of strings
        # headers are in row 1
        from openpyxl.utils import get_column_letter
        
        # Specific columns to widen (2x)
        double_width_cols = {
            "產品類型", "操作系統", "行銷目標", "計費模式", 
            "受眾操作系統", "最高系統版本", "深度轉換目標", "轉換價值", "轉化目標"
        }

        # We process header first
        column_widths = {}
        for i, col_name in enumerate(columns):
            # 1.5 factor for bold font and some padding
            width = len(str(col_name)) * 1.5 + 2
            
            # Check if this column needs doubling (based on name)
            # Handle duplicate names logic: The name in 'columns' list is what we check.
            if col_name in double_width_cols:
                # User asked for "column width add 2 times" (width * 2 presumably, or increase significantly)
                # I will store a multiplier.
                pass # Applied below
                
            column_widths[i+1] = width

        # Process data rows
        # Since dataframe can be large, iterating again might be slow but OK for this scale (~few k rows).
        # We can iterate the dataframe 'rows' list we constructed earlier 'rows' variable, which matches Excel except sparse logic.
        # Actually 'rows' variable in the loops above handles the sparse logic logic?
        # Yes, 'rows' list contains the actual data to be written.
        # Iterate 'rows' (which are lists of values)
        for r in rows:
            for i, val in enumerate(r):
                if val:
                    val_len = len(str(val))
                    # Adjust factor for chinese characters (width ~2) vs english (width ~1)
                    # Simple heuristic: len * 1.3 + padding
                    # Or count bytes?
                    # Let's use simple length * 1.8 for safety to fit font 12
                    curr_w = val_len * 1.8 
                    if curr_w > column_widths[i+1]:
                        column_widths[i+1] = curr_w
        
        # Set widths
        for col_idx, width in column_widths.items():
            # Apply doubling for specific columns
            # Get column name from our list (0-based)
            if col_idx <= len(columns):
                col_name = columns[col_idx-1]
                if col_name in double_width_cols:
                    width = width * 2
            
            # Cap width to avoid overly wide columns (e.g. long URL)
            final_width = min(width, 100) # Increased cap for doubled columns
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = final_width
                    
        # --- Data Validation ---
        # Max row for validation
        dv_max = max_r + 100
        
        # Helper to add DV
        def _add_dv(col_letter, options_list, prompt_title=""):
            # Ensure options are strings and comma separated within double quotes
            quoted_opts = [f'{o}' for o in options_list]
            formula = f'"{",".join(quoted_opts)}"'
            
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.error = '請從下拉選單中選擇有效的值'
            dv.errorTitle = '輸入錯誤'
            dv.prompt = '請從選單中選擇'
            dv.promptTitle = prompt_title
            
            worksheet.add_data_validation(dv)
            dv.add(f'{col_letter}2:{col_letter}{dv_max}')

        # 1. Status Columns (C, L, AS)
        _add_dv("C", ["開啟", "關閉"], "廣告活動狀態")
        _add_dv("L", ["開啟", "關閉"], "廣告群組狀態")

        # 2. Platform (AE) & Browser (AG)
        # Note: Columns indices:
        # "受眾操作系統" -> Index 30 -> 31st col -> AE
        # "瀏覽器" -> Index 32 -> 33rd col -> AG
        _add_dv("AE", list(PLATFORM_MAP.values()), "受眾操作系統")
        _add_dv("AG", list(BROWSER_MAP.values()), "瀏覽器")
        
        # Creative Status: 
        # Index 44 -> AR (0-based 44 -> 45th col -> 45-26=19 -> S)
        # Wait, let's re-verify Col 44.
        # A=1... Z=26. AA=27.
        # Col 1 = A.
        # Col 45. 45-26 = 19. 19th char is S. (A=1... S=19).
        # So AR is 18?
        # R is 18th letter. AR is 26+18 = 44.
        # So AR is column 44.
        # My Columns list has "廣告文案狀態" at index 44.
        # Python list index 44 is the 45th element.
        # So it is Column 45.
        # Column 45 is AS. (26+19=45). 
        # Previous logic: 40 -> AN. (26+14=40). 39 index is 40th col -> AN. Correct.
        # 44 index is 45th col -> AS. 
        # So Creative Status WAS "AS". My previous code had "AS" then I changed to "AR"?
        # Let's check listing again.
        # ... AP(41), AQ(42), AR(43), AS(44)?
        # Index 39=AN.
        # Index 40=AO.
        # Index 41=AP. (廣告文案名稱)
        # Index 42=AQ. (廣告文案ID)
        # Index 43=AR. (廣告文案狀態) -> THIS IS IT.
        # Wait, check columns list provided in code:
        # ... "AI語意擴充關鍵字"(40)
        # "廣告文案名稱"(41)
        # "廣告文案ID"(42)
        # "廣告文案狀態"(43)
        # So "廣告文案狀態" is index 43.
        # Index 43 is the 44th column.
        # 44th column -> 44-26 = 18 -> R. 
        # So AR is indeed correct for Index 43.
        # My apologies, I need to be super precise.
        
        # Col Index 43 (0-based) -> Excel Col 44 -> AR.
        _add_dv("AR", ["開啟", "關閉", "已過期"], "廣告文案狀態")

        # 2. 產品類型 (Col 5 -> F)
        _add_dv("F", ["web", "app"], "產品類型")

        # 3. 操作系統 (APP) (Col 7 -> H)
        _add_dv("H", ["iOS", "Android", "Others"], "操作系統")
        
        # 4. 行銷目標 (Col 15 -> P) (Index 15 -> 16th col -> P)
        _add_dv("P", ["品牌知名度", "電商網上購買", "增加網站流量", "開發潛在客戶", "網站互動"], "行銷目標")

        # 5. 計費模式 (Col 16 -> Q) (Index 16 -> 17th col -> Q)
        _add_dv("Q", ["CPM", "CPC"], "計費模式")

        # 6. 深度轉換目標 (Col 19 -> T) (Index 19 -> 20th col -> T)
        _add_dv("T", ["帳戶預設設定", "所有轉換", "指定轉換目標"], "深度轉換目標")

        # 6.5 轉化目標 (Col 21 -> V)
        _add_dv("V", ["點擊數", "網頁瀏覽", "完成註冊", "搜尋", "收藏", "加入購物車", "開始結帳", "完成結帳"], "轉化目標")

        # 7. 地理位置 (Col 26 -> AA) (Index 26 -> 27th col -> AA)
        _add_dv("AA", ["包含", "不包含"], "地理位置")
        
        # 8. 受眾操作系統 (Col 30 -> AE)
        _add_dv("AE", ["iOS", "Android", "Others"], "受眾操作系統")
        
        # 9. 投放興趣選項 (Col 35 -> AJ) (Index 35 -> 36th col -> AJ)
        _add_dv("AJ", ["包含", "不包含"], "投放興趣選項")
        
        # 10. AI語意擴充選項 (Col 39 -> AN) (Index 39 -> 40th col -> AN)
        _add_dv("AN", ["包含", "不包含"], "AI語意擴充選項")
        
    return output.getvalue()
